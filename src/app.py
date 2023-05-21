import openpyxl
from tabulate import tabulate

# Define variable to load the dataframe
# excel_dataframe = openpyxl.load_workbook("personas.xlsx")

def buscarLibros(titulo):
    if titulo == "IA":
        excel_dataframe = openpyxl.load_workbook("libros.xlsx")
        return leerLista(excel_dataframe)
    elif titulo == "PROGRAMACION":
        excel_dataframe = openpyxl.load_workbook("programacion.xlsx")
        return leerLista(excel_dataframe)
    else:
        return "No se encontraron libros para " + titulo

def leerLista(excel_dataframe):
    # Define variable to read sheet
    dataframe = excel_dataframe.active

    data = []

    # Iterate the loop to read the cell values
    for row in range(1, dataframe.max_row):
        _row = [row,]

        for col in dataframe.iter_cols(0, dataframe.max_column):
            _row.append(col[row].value)

        data.append(_row)

    headers = ["#", "titulo", "autor", "keyword", "link"]
    headers_align = (("left",) * 0 )

    return tabulate(data, headers=headers, colalign=headers_align)
